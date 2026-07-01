import os
import pdfplumber
import pandas as pd
import re
from openpyxl import Workbook


def extract_strength_triplet(text, aliases, exclude_aliases=None):
    """
    从 PDF 文本中按行提取三列强度值。
    支持单值和“a/b”这类双值文本，抽不到时返回空字符串。
    """
    number_pattern = r'[-+]?\d+(?:\.\d+)?(?:\s*/\s*[-+]?\d+(?:\.\d+)?)?'
    normalized_aliases = [re.sub(r'\s+', '', alias) for alias in aliases]
    normalized_exclude_aliases = [re.sub(r'\s+', '', alias) for alias in (exclude_aliases or [])]

    for raw_line in text.splitlines():
        line = raw_line.strip()
        compact_line = re.sub(r'\s+', '', line)
        if not line or not any(alias in compact_line for alias in normalized_aliases):
            continue
        if any(alias in compact_line for alias in normalized_exclude_aliases):
            continue

        values = re.findall(number_pattern, line)
        if len(values) >= 3:
            return values[0], values[1], values[2]

    return '', '', ''


def extract_value_pairs_near_label(text, label, expected_count, end_labels=None):
    """
    以参数标签为锚点截取局部窗口，再提取顺序数对。
    同时兼容“每组数对各带标签”和“一个标签后连续多组数对”的 PDF 文本。
    """
    anchor_match = re.search(rf'\[{re.escape(label)}\]', text, re.IGNORECASE)
    if not anchor_match:
        return []

    start_index = max(0, anchor_match.start() - 200)
    end_index = len(text)
    for end_label in end_labels or []:
        end_match = re.search(re.escape(end_label), text[anchor_match.end():], re.IGNORECASE)
        if end_match:
            candidate_end = anchor_match.end() + end_match.start()
            if candidate_end < end_index:
                end_index = candidate_end

    section_text = text[anchor_match.start():end_index]

    pair_pattern = re.compile(
        r'([-+]?\d+(?:\.\d+)?)\s*/\s*([-+]?\d+(?:\.\d+)?)',
        re.IGNORECASE | re.DOTALL
    )
    matches = pair_pattern.findall(section_text)
    return matches[:expected_count]


def create_strength_summary_df(sun_gear, planet_gear, ring_gear):
    """
    构造强度汇总 sheet，按“指标为行、齿轮类型为列”输出。
    """
    strength_rows = [
        ('齿根应力安全系数 SF', '齿根应力安全系数_SF'),
        ('接触应力安全系数 SHBD', '接触应力安全系数_SH'),
        ('总重合度 εγ', '总重合度_εγ')
    ]

    summary_data = {
        '强度指标': [],
        '太阳轮': [],
        '行星轮': [],
        '齿圈': []
    }

    for display_name, data_key in strength_rows:
        summary_data['强度指标'].append(display_name)
        summary_data['太阳轮'].append(sun_gear.get(data_key, ''))
        summary_data['行星轮'].append(planet_gear.get(data_key, ''))
        summary_data['齿圈'].append(ring_gear.get(data_key, ''))

    return pd.DataFrame(summary_data)

def extract_gear_parameters_from_pdf(pdf_path):
    """
    从PDF文件中提取齿轮参数
    """
    # 读取PDF文本
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text += (page.extract_text() or "") + "\n"
    
    # 初始化三个齿轮的数据字典
    sun_gear = {}
    planet_gear = {}
    ring_gear = {}
    
    # 提取基本参数
    patterns = {
        '齿数': r'齿数.*?\[z\].*?(\d+)\s+(\d+)\s+([-\d]+)',
        '法向模数': r'法向模数.*?\[mn\].*?([\d.]+)',
        '压力角': r'法向压力角.*?\[αn\].*?([\d.]+)',
        '螺旋角': r'分度圆上的螺旋角.*?\[β\].*?([\d]+)',
        '螺旋方向': r'螺旋线方向.*?[\u4e00-\u9fa5]+啮合',
        '齿顶高系数': r'基准齿廓齿顶高.*?\[haP\*\].*?([\d.]+)',
        '齿根高系数': r'基准齿廓齿根高.*?\[hfP\*\].*?([\d.]+)\s+([\d.]+)\s+([\d.]+)',
        '齿廓变位系数': r'齿廓变位系数.*?\[x\].*?([\d.-]+)\s+([\d.-]+)\s+([-\d.-]+)',
        '齿根圆直径': r'齿根圆直径.*?\[df\].*?([\d.]+)\s+([\d.]+)\s+([-\d.]+)',
        '齿顶圆直径': r'齿顶圆直径.*?\[da\].*?([\d.]+)\s+([\d.]+)\s+([-\d.]+)',
        '渐开线起始圆': r'齿根成形圆直径.*?\[dFf\].*?([\d.]+)\s+([\d.]+)\s+([-\d.]+)',
        '齿根圆角系数': r'基准齿廓齿根半径.*?\[ρfP\*\].*?([\d.]+)\s+([\d.]+)\s+([\d.]+)',
        '齿宽': r'齿宽.*?\[b\].*?([\d.]+)\s+([\d.]+)\s+([\d.]+)',
        '中心距': r'中心距.*?\[a\].*?([\d.]+)',
        '跨齿数': r'跨齿数.*?\[k\].*?([\d.]+)\s+([\d.]+)\s+([-\d.]+)',
        '量棒直径': r'有效量规直径.*?\[DMeff\].*?([\d.]+)\s+([\d.]+)\s+([\d.]+)',
        '单个齿距偏差': r'单个齿距偏差的公差.*?\[fpt\].*?([\d.]+)\s+([\d.]+)\s+([\d.]+)',
        '齿距累计总偏差': r'齿距累积总偏差的公差.*?\[FPT\].*?([\d.]+)\s+([\d.]+)\s+([\d.]+)',
        '齿廓总偏差': r'齿廓总偏差的公差.*?\[FαT\].*?([\d.]+)\s+([\d.]+)\s+([\d.]+)',
        '螺旋线总偏差': r'螺旋线总偏差的公差.*?\[FβT\].*?([\d.]+)\s+([\d.]+)\s+([\d.]+)',
        '径向跳动偏差': r'径跳偏差的公差.*?\[FrT\].*?([\d.]+)\s+([\d.]+)\s+([\d.]+)'
    }
    # 提取数据
    for param_name, pattern in patterns.items():
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            if param_name in ['齿数', '齿廓变位系数', '齿根圆直径', '齿顶圆直径', '渐开线起始圆', '齿根高系数','齿根圆角系数', '齿宽',
                            '跨齿数', '量棒直径', '单个齿距偏差', '齿距累计总偏差',
                            '齿廓总偏差', '螺旋线总偏差', '径向跳动偏差']:
                groups = match.groups()
                if len(groups) >= 3:
                    sun_gear[param_name] = groups[0]
                    planet_gear[param_name] = groups[1]
                    ring_gear[param_name] = groups[2]
            elif param_name in ['法向模数', '压力角', '螺旋角', '齿顶高系数',  
                               '中心距']:
                sun_gear[param_name] = match.group(1)
                planet_gear[param_name] = match.group(1)
                ring_gear[param_name] = match.group(1)
            elif param_name == '螺旋方向':
                sun_gear[param_name] = '直齿'
                planet_gear[param_name] = '直齿'
                ring_gear[param_name] = '直齿'
    
    # 处理行星轮数量
    planet_count_match = re.search(r'齿轮数量.*?\[p\].*?1\s+(\d+)\s+1', text)
    if planet_count_match:
        planet_gear['数量'] = planet_count_match.group(1)

    # 处理强度汇总指标，按 PDF 表格行抽取三列数据
    strength_mapping = {
        '齿根应力安全系数_SF': {
            'aliases': ['齿根应力安全系数', 'SF=σFG/σF'],
            'exclude_aliases': ['SFmin', '目标安全系数', '齿根目标安全系数']
        },
        '接触应力安全系数_SH': {
            'aliases': ['单对齿啮合齿面接触应力安全系数', 'SHBD=σHG/σHBD', '[SHBD'],
            'exclude_aliases': ['SHmin', 'SHw', '(SHBD)²', '目标安全系数', '齿面目标安全系数']
        }
    }
    for field_name, rule in strength_mapping.items():
        sun_value, planet_value, ring_value = extract_strength_triplet(
            text,
            rule['aliases'],
            rule.get('exclude_aliases')
        )
        if sun_value or planet_value or ring_value:
            sun_gear[field_name] = sun_value
            planet_gear[field_name] = planet_value
            ring_gear[field_name] = ring_value

    # 总重合度是啮合副指标。样例中两值分别对应太阳轮-行星轮和行星轮-齿圈。
    total_contact_ratio_match = re.search(r'总重合度.*?\[εγ\].*?([-+]?\d+(?:\.\d+)?)\s+([-+]?\d+(?:\.\d+)?)', text)
    if total_contact_ratio_match:
        first_mesh_value = total_contact_ratio_match.group(1)
        second_mesh_value = total_contact_ratio_match.group(2)
        sun_gear['总重合度_εγ'] = first_mesh_value
        planet_gear['总重合度_εγ'] = f'{first_mesh_value} / {second_mesh_value}'
        ring_gear['总重合度_εγ'] = second_mesh_value
    
    # 围绕首个 Wk.e/i 局部提取公法线，避免章节标题 OCR 和全文顺序扫描带来的错位。
    wk_pairs = extract_value_pairs_near_label(text, 'Wk.e/i', 3, ['[dMWk.m]', '[MrK.e/i]'])
    if len(wk_pairs) >= 1:
        sun_gear['公法线长度_Wmax'] = wk_pairs[0][0]
        sun_gear['公法线长度_Wmin'] = wk_pairs[0][1]
    if len(wk_pairs) >= 2:
        planet_gear['公法线长度_Wmax'] = wk_pairs[1][0]
        planet_gear['公法线长度_Wmin'] = wk_pairs[1][1]
    if len(wk_pairs) >= 3:
        ring_gear['公法线长度_Wmax'] = wk_pairs[2][0]
        ring_gear['公法线长度_Wmin'] = wk_pairs[2][1]

    # 处理 HAC 高度处的齿厚偏差对 [sc.e/i]
    # PDF 中通常会连续出现三行 [sc.e/i]，分别对应太阳轮、行星轮、齿圈
    sc_matches = re.findall(r'\[sc\.e/i\].*?([\d.]+)\s*/\s*([\d.]+)', text, re.IGNORECASE | re.DOTALL)
    if len(sc_matches) >= 3:
        sun_gear['hac高度处齿厚_e'] = sc_matches[0][0]
        sun_gear['hac高度处齿厚_i'] = sc_matches[0][1]
        planet_gear['hac高度处齿厚_e'] = sc_matches[1][0]
        planet_gear['hac高度处齿厚_i'] = sc_matches[1][1]
        ring_gear['hac高度处齿厚_e'] = sc_matches[2][0]
        ring_gear['hac高度处齿厚_i'] = sc_matches[2][1]

    # 处理齿顶圆公差带
    da_pattern = r'\[da\.e/i\].*?([\d.]+)\s+/\s+([\d.]+)'
    sun_da_match = re.search(da_pattern, text)
    if sun_da_match:
        try:
            sun_da_max = float(sun_da_match.group(1))
            sun_da_min = float(sun_da_match.group(2))
            sun_gear['齿顶公差范围'] = sun_da_max - sun_da_min
        except:
            sun_gear['齿顶公差范围'] = 0
    else:
        da_pattern = r'\[da\.e/i\].*?([\d.]+)\s+/([\d.]+)'
        sun_da_match = re.search(da_pattern, text)
        try:
            sun_da_max = float(sun_da_match.group(1))
            sun_da_min = float(sun_da_match.group(2))
            sun_gear['齿顶公差范围'] = sun_da_max - sun_da_min
        except:
            sun_gear['齿顶公差范围'] = 0
    
    if sun_da_match:
        remaining_text = text[sun_da_match.end():]
        planet_pattern = r'([\d.]+)\s+/([\d.]+)'
        planet_da_match = re.search(planet_pattern, remaining_text)
        if planet_da_match:
            try:
                planet_da_max = float(planet_da_match.group(1))
                planet_da_min = float(planet_da_match.group(2))
                planet_gear['齿顶公差范围'] = planet_da_max - planet_da_min
            except:
                planet_gear['齿顶公差范围'] = 0
        if planet_da_match:
            remaining_text = text[planet_da_match.end():]
            planet_w_pattern = r'\[Ada\.e/i\].*?([\d.]+)'
            ring_da_match = re.search(planet_w_pattern, remaining_text)
            remaining_text1 = remaining_text[ring_da_match.end():]
            ring_da_match1 = re.search(planet_w_pattern, remaining_text1)
            remaining_text2 = remaining_text1[ring_da_match1.end():]
            ring_da_match2 = re.search(planet_w_pattern, remaining_text2)
            remaining_text2 = remaining_text2[ring_da_match1.end():]
            
            if ring_da_match2:
                try:
                    ring_da_max = float(ring_da_match2.group(1))
                    ring_gear['齿顶公差范围'] = ring_da_max
                except:
                    ring_gear['齿顶公差范围'] = 0

    # 处理跨棒距（齿圈）- 径向二针跨球距
    md_pattern = r'径向二针跨球距.*?\[MdK\.e/i\].*?([\d.]+)\s+/\s+([\d.]+)'
    md_match = re.search(md_pattern, text)
    if not md_match:
        md_pattern = r'径向二针跨球距.*?\[MdK\.e/i\].*?([\d.]+)\s+/([\d.]+)'
        md_match = re.search(md_pattern, text)
        
    md_match1=None    

    if md_match:
        remaining_text = text[md_match.end():]
        md_pattern = r'\[MdK\.e/i\].*?([\d.]+)'
        md_match = re.search(md_pattern, remaining_text)
        if float(md_match.group(1))>=10:
            md_pattern = r'([\d.]+)\s+/([\d.]+)'
            md_match1 = re.search(md_pattern, remaining_text)
        elif float(md_match.group(1))>0 and float(md_match.group(1))<10:
            md_pattern = r'([\d.]+)\s+/\s+([\d.]+)'
            md_match1 = re.search(md_pattern, remaining_text)
        else:
            md_match1=None

    if md_match1:
        remaining_text1 = remaining_text[md_match1.end():]
        md_pattern = r'\[MdK\.e/i\].*?([\d.]+)'
        md_match = re.search(md_pattern, remaining_text1)
        if float(md_match.group(1)) >= 10:
            md_pattern = r'([\d.]+)\s+/([\d.]+)'
            md_match2 = re.search(md_pattern, remaining_text1)
        elif float(md_match.group(1)) > 0 and float(md_match.group(1)) < 10:
            md_pattern = r'([\d.]+)\s+/\s+([\d.]+)'
            md_match2 = re.search(md_pattern, remaining_text1)
        else:
            md_match2 = None

    if md_match2:
        ring_gear['跨棒距_max'] = md_match2.group(1)
        ring_gear['跨棒距_min'] = md_match2.group(2)

    # 如果没有找到齿廓变位系数，尝试从产形齿廓变位系数中提取（作为备选）
    if '齿廓变位系数' not in sun_gear:
        backup_pattern = r'产形齿廓变位系数.*?\[xE e/i\].*?([\d.-]+).*?([\d.-]+).*?([\d.-]+)'
        backup_match = re.search(backup_pattern, text)
        if backup_match:
            sun_gear['齿廓变位系数'] = backup_match.group(1)
            planet_gear['齿廓变位系数'] = backup_match.group(2)
            ring_gear['齿廓变位系数'] = backup_match.group(3)
    
    # 计算顶隙系数：齿根高系数 - 齿顶高系数
    if '齿根高系数' in sun_gear and '齿顶高系数' in sun_gear:
        try:
            hf = float(sun_gear['齿根高系数'])
            ha = float(sun_gear['齿顶高系数'])
            c_value = hf - ha
            sun_gear['顶隙系数'] = f"{c_value:.2f}"
            hf = float(planet_gear['齿根高系数'])
            ha = float(planet_gear['齿顶高系数'])
            c_value = hf - ha
            planet_gear['顶隙系数'] = f"{c_value:.2f}"
            hf = float(ring_gear['齿根高系数'])
            ha = float(ring_gear['齿顶高系数'])
            c_value = hf - ha
            ring_gear['顶隙系数'] = f"{c_value:.2f}"
        except ValueError:
            sun_gear['顶隙系数'] = "0.25"
            planet_gear['顶隙系数'] = "0.25"
            ring_gear['顶隙系数'] = "0.25"
    else:
        # 如果无法计算，使用默认值
        sun_gear['顶隙系数'] = "0.25"
        planet_gear['顶隙系数'] = "0.25"
        ring_gear['顶隙系数'] = "0.25"
    
    # 设置精度等级
    sun_gear['精度等级'] = "ISO1328"
    planet_gear['精度等级'] = "ISO1328"
    ring_gear['精度等级'] = "ISO1328"
    
    # 格式化数值
    def format_value(value, param_name):
        if not value:
            return ""
        if param_name in ['中心距', '齿根圆直径', '渐开线起始圆']:
            try:
                return f"{float(value):.4f}"
            except:
                return value
        # HAC 高度处齿厚偏差对按小数 4 位格式化
        if param_name in ['hac高度处齿厚', 'hac高度处齿厚_e', 'hac高度处齿厚_i']:
            try:
                return f"{float(value):.4f}"
            except:
                return value
        if param_name in ['压力角', '螺旋角']:
            try:
                return f"{float(value):.1f}°"
            except:
                return value
        if param_name in ['齿数','跨齿数']:
            try:
                return f"{int(value.split('.')[0])}"
            except:
                return value        
        if param_name in ['单个齿距偏差']:
            try:
                return "±"+f"{str(value)}"
            except:
                return value
        return value
    
    # 构建标准化的数据结构
    def create_gear_data(gear_dict, gear_type):
        # 齿轮参数部分表格内容
        gear_params = {
            '参数名称': [],
            '符号': [],
            '数值': []
        }

        # 主参数映射（去除齿顶圆直径、数量）
        param_mapping = {
            '齿数': ('Z', '齿数', 'gear'),
            '法向模数': ('mn', '法向模数', 'gear'),
            '压力角': ('α', '压力角', 'gear'),
            '螺旋角': ('β', '螺旋角', 'gear'),
            '螺旋方向': ('', '螺旋方向', 'gear'),
            '齿顶高系数': ('ha*', '齿顶高系数', 'gear'),
            '顶隙系数': ('C*', '顶隙系数', 'gear'),
            '齿廓变位系数': ('x', '径向变位系数', 'gear'),
            '齿根圆直径': ('df', '齿根圆直径', 'gear'),
            '渐开线起始圆': ('dFf', '渐开线起始圆', 'gear'),
            '齿根圆角系数': ('rhofP*', '齿根圆角系数', 'gear'),
            '中心距': ('a', '中心距', 'gear'),
            '相配齿轮图号': ('', '相配齿轮图号', 'gear'),
            '相配齿轮齿数': ('', '相配齿轮齿数', 'gear'),
            '精度等级': ('6', '精度等级', 'gear'),
            '跨齿数': ('k', '跨齿数', 'gear'),
            '公法线长度_Wmax': ('Wmax', '公法线长度', 'gear'),
            '公法线长度_Wmin': ('Wmin', '', 'gear'),
            '量棒直径': ('DM', '量棒直径', 'gear'),
            '跨棒距_max': ('Mmax', '跨棒距', 'gear'),
            '跨棒距_min': ('Mmin', '', 'gear')
        }

        # 其它参数单独映射
        others_mapping = {
            '齿顶圆直径': ('da', '齿顶圆直径', 'gear'),
            'hac高度处齿厚_e': ('sc.e', '弦齿厚', 'gear'),
            'hac高度处齿厚_i': ('sc.i', '', 'gear'),
            '齿宽': ('b', '齿宽', 'gear'),
            '数量': ('', '数量', 'gear')
        }
        
        
        # 精度参数映射
        accuracy_mapping = {
            '单个齿距偏差': ('±fpt', '单个齿距偏差', 'accuracy'),
            '齿距累计总偏差': ('Fp', '齿距累计总偏差', 'accuracy'),
            '齿廓总偏差': ('Fɑ', '齿廓总偏差', 'accuracy'),
            '螺旋线总偏差': ('Fβ', '螺旋线总偏差', 'accuracy'),
            '径向跳动偏差': ('Fr', '径向跳动偏差', 'accuracy'),
            '齿顶公差范围': ('', '齿顶公差范围', 'accuracy')
        }
        
        # 先处理主参数
        for param_key, (symbol, display_name, param_type) in param_mapping.items():
            if param_type == 'gear':
                # 太阳轮和行星轮没有量棒直径
                if gear_type in ['太阳轮', '行星轮'] and param_key == '量棒直径':
                    continue
                if param_key in gear_dict:
                    formatted_value = format_value(gear_dict[param_key], display_name)
                    gear_params['参数名称'].append(display_name)
                    gear_params['符号'].append(symbol)
                    gear_params['数值'].append(formatted_value)
                elif display_name in ['相配齿轮图号', '相配齿轮齿数']:
                    # 保留空行
                    gear_params['参数名称'].append(display_name)
                    gear_params['符号'].append(symbol)
                    gear_params['数值'].append('')

        # 精度参数同原逻辑
        
        # 添加齿轮精度标题
        gear_params['参数名称'].append('齿轮精度')
        gear_params['符号'].append('')
        gear_params['数值'].append('')

        # 处理齿轮精度参数
        for param_key, (symbol, display_name, param_type) in accuracy_mapping.items():
            if param_type == 'accuracy' and param_key in gear_dict:
                formatted_value = format_value(gear_dict[param_key], display_name)
                gear_params['参数名称'].append(display_name)
                gear_params['符号'].append(symbol)
                gear_params['数值'].append(formatted_value)

        # others_mapping内容插入，顺序为齿顶圆直径、sc.e、sc.i、齿宽、数量
        for key in ['齿顶圆直径', 'hac高度处齿厚_e', 'hac高度处齿厚_i', '齿宽', '数量']:
            symbol, display_name, _ = others_mapping[key]
            if key in gear_dict:
                # 这里使用参数键做格式化判定，避免空显示名（如 sc.i 第二行）丢失格式规则
                formatted_value = format_value(gear_dict[key], key)
                gear_params['参数名称'].append(display_name)
                gear_params['符号'].append(symbol)
                gear_params['数值'].append(formatted_value)

        return gear_params
    
    # 特殊处理：齿圈没有跨齿数，太阳轮和行星轮没有跨棒距和量棒直径
    if '跨齿数' in ring_gear:
        del ring_gear['跨齿数']
    if '跨棒距_max' in sun_gear:
        del sun_gear['跨棒距_max']
        del sun_gear['跨棒距_min']
    if '跨棒距_max' in planet_gear:
        del planet_gear['跨棒距_max']
        del planet_gear['跨棒距_min']
    if '公法线长度_Wmax' in ring_gear:
        del ring_gear['公法线长度_Wmax']
        del ring_gear['公法线长度_Wmin']
    if '量棒直径' in sun_gear:
        del sun_gear['量棒直径']
    if '量棒直径' in planet_gear:
        del planet_gear['量棒直径']
    
    sun_data = create_gear_data(sun_gear, '太阳轮')
    planet_data = create_gear_data(planet_gear, '行星轮')
    ring_data = create_gear_data(ring_gear, '齿圈')
    
    # 修改列名为指定的名称
    def rename_columns(data, sheet_name):
        df = pd.DataFrame(data)
        if sheet_name == '齿圈':
            df = df.rename(columns={'参数名称': '齿圈参数'})
        else:
            df = df.rename(columns={'参数名称': '齿轮参数'})
        return df
    
    sun_df = rename_columns(sun_data, '太阳轮')
    planet_df = rename_columns(planet_data, '行星轮')
    ring_df = rename_columns(ring_data, '齿圈')
    
    strength_df = create_strength_summary_df(sun_gear, planet_gear, ring_gear)

    return sun_df, planet_df, ring_df, strength_df

def process_all_pdfs(input_dir="./input", output_dir="./excel"):
    """
    处理input文件夹中的所有PDF文件
    """
    if not os.path.exists(input_dir):
        print(f"输入文件夹 {input_dir} 不存在")
        return
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # 获取所有PDF文件
    pdf_files = [f for f in os.listdir(input_dir) if f.lower().endswith('.pdf') and f.lower().startswith('gear')]
    
    if not pdf_files:
        print(f"在 {input_dir} 文件夹中未找到PDF文件")
        return
    
    for pdf_file in pdf_files:
        pdf_path = os.path.join(input_dir, pdf_file)
        excel_name = os.path.splitext(pdf_file)[0] + '.xlsx'
        output_path = os.path.join(output_dir, excel_name)
        
        try:
            print(f"正在处理: {pdf_file}")
            sun_df, planet_df, ring_df, strength_df = extract_gear_parameters_from_pdf(pdf_path)
            
            # 保存为Excel文件
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                sun_df.to_excel(writer, sheet_name='太阳轮', index=False)
                planet_df.to_excel(writer, sheet_name='行星轮', index=False)
                ring_df.to_excel(writer, sheet_name='齿圈', index=False)
                strength_df.to_excel(writer, sheet_name='强度汇总', index=False)
            # ========== 自动列宽与居中设置，详细中文注释 ==========
            from openpyxl.styles import Alignment
            from openpyxl.utils import get_column_letter
            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # 1. 遍历所有单元格，设置水平和垂直居中
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                # 2. 获取表头（第一行），用于判断每一列的类型
                header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for idx, col in enumerate(ws.columns, 0):
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    # 3. 遍历本列所有单元格，取最大内容宽度
                    for cell in col:
                        try:
                            value = str(cell.value) if cell.value is not None else ''
                            if len(value) > max_length:
                                max_length = len(value)
                                # print(f"列 {col_letter} 当前最大内容宽度: {max_length} (单元格 {cell.coordinate} 内容: '{value}')")
                        except Exception:
                            pass
                    # 4. 判断列名，决定列宽加宽策略
                    if idx < len(header):
                        col_name = header[idx]
                        # “参数名称”类列名，列宽=最大内容宽度+4
                        if col_name in ['参数名称', '齿轮参数', '齿圈参数', '参数', '参数名', '参数项']:
                            ws.column_dimensions[col_letter].width = max_length + 7
                        # “数值”类列名，列宽=最大内容宽度+2
                        elif col_name in ['数值', '值', '参数值']:
                            ws.column_dimensions[col_letter].width = 9
                        else:
                            # 其它列默认+2
                            ws.column_dimensions[col_letter].width = max_length + 2
                    else:
                        ws.column_dimensions[col_letter].width = max_length + 2
                # ======= 列宽计算完成后再执行合并操作，避免合并影响列宽判断 =======
                # 合并顶行大标题（方案A）：如果首列为分组标题，则合并 A1:C1 并居中
                try:
                    first_val = ws.cell(1, 1).value
                    if first_val in ['齿轮参数', '齿圈参数', '参数名称']:
                        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
                        ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(1, 2).value = None
                        ws.cell(1, 3).value = None
                except Exception:
                    pass

                # 遍历数据区，遇到分组标题如 '齿轮精度' 时合并该行 A-C 并居中
                try:
                    for r in range(1, ws.max_row + 1):
                        v = ws.cell(r, 1).value
                        if v == '齿轮精度':
                            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
                            ws.cell(r, 1).alignment = Alignment(horizontal='center', vertical='center')
                            ws.cell(r, 2).value = None
                            ws.cell(r, 3).value = None
                except Exception:
                    pass
            wb.save(output_path)
            # ========== 以上代码实现了所有单元格内容居中和自动调整列宽 ==========
            print(f"成功输出: {excel_name}")
        except Exception as e:
            print(f"处理文件 {pdf_file} 时出错: {str(e)}")

# 执行处理
if __name__ == "__main__":
    process_all_pdfs()