import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import shutil
import re
import logging
from datetime import datetime
import sys

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None

# 配置日志，便于排查监听与升级过程中的问题
LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "update_file.log")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class DrawingManager:
    def __init__(self, root, default_path="D:/样品/项目/图纸"):
        self.root = root
        self.root.title("图纸/文档版本管理系统 v1.4")
        self.root.geometry("980x680")

        self.base_path = tk.StringVar(value=default_path)
        self.current_selection = tk.StringVar()
        self.version_input = tk.StringVar()
        self.remark_input = tk.StringVar()
        self.monitor_status = tk.StringVar(value="自动监听：运行中（每3秒扫描）")

        # 自动监听状态：记录当前处于“修改中”的图纸
        self.lock_poll_ms = 3000
        self.active_dwg_locks = {}
        self.scan_job = None

        self.setup_ui()
        self.refresh_tree()
        self.start_monitoring()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def setup_ui(self):
        top_frame = ttk.Frame(self.root, padding=10)
        top_frame.pack(fill="x")
        ttk.Label(top_frame, text="根目录:").pack(side="left")
        ttk.Entry(top_frame, textvariable=self.base_path, width=70).pack(side="left", padx=5)
        ttk.Button(top_frame, text="浏览", command=self.browse_path).pack(side="left")
        ttk.Button(top_frame, text="刷新", command=self.refresh_tree).pack(side="left", padx=5)
        ttk.Label(top_frame, textvariable=self.monitor_status).pack(side="left", padx=10)

        tree_frame = ttk.Frame(self.root, padding=10)
        tree_frame.pack(fill="both", expand=True)
        self.tree = ttk.Treeview(tree_frame, columns=("date", "type"), show="tree headings")
        self.tree.heading("#0", text="名称")
        self.tree.heading("date", text="修改日期")
        self.tree.heading("type", text="类型")
        self.tree.column("#0", width=400)
        self.tree.pack(side="left", fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        self.tree.bind("<Button-3>", self.show_context_menu)

        # 右键菜单：支持打开文件和打开所在目录
        self.tree_menu = tk.Menu(self.root, tearoff=0)
        self.tree_menu.add_command(label="打开", command=self.open_selected_item)
        self.tree_menu.add_command(label="打开所在目录", command=self.open_selected_folder)

        op_frame = ttk.LabelFrame(self.root, text="操作面板", padding=15)
        op_frame.pack(fill="x", padx=10, pady=10)
        ttk.Label(op_frame, text="版本号:").grid(row=0, column=0, sticky="w")
        ttk.Entry(op_frame, textvariable=self.version_input, width=30).grid(row=0, column=1, padx=10)
        ttk.Label(op_frame, text="备注:").grid(row=0, column=2, sticky="w")
        ttk.Entry(op_frame, textvariable=self.remark_input, width=36).grid(row=0, column=3, padx=10)
        
        btn_box = ttk.Frame(op_frame)
        btn_box.grid(row=1, column=0, columnspan=4, pady=15)
        ttk.Button(btn_box, text="📁 升级文件夹版本", command=self.upgrade_folder).pack(side="left", padx=10)
        ttk.Button(btn_box, text="📂 新建版本文件夹", command=self.create_version_folder).pack(side="left", padx=10)
        ttk.Button(btn_box, text="📑 升级图纸/PDF文件", command=self.upgrade_file).pack(side="left", padx=10)
        ttk.Button(btn_box, text="📑 批量升级文件", command=self.bulk_upgrade_files).pack(side="left", padx=5)

        self.root.bind("<Control-d>", self.upgrade_folder)
        self.root.bind("<Control-f>", self.upgrade_file)
        self.root.bind("<Delete>", self.delete_items)
        self.root.bind("<Control-Shift-F>", self.bulk_upgrade_files)
        delete_btn = ttk.Button(btn_box, text="🗑️ 删除选中项 (Del)", command=self.delete_items)
        delete_btn.pack(side="left", padx=20)
        
        # 新增：删除旧版本按钮
        ttk.Button(btn_box, text="🗑️ 删除旧版本，只保留最新", command=self.delete_old_versions).pack(side="left", padx=5)

    def start_monitoring(self):
        """启动目录监听，每3秒扫描一次 .dwl/.dwl2 锁文件"""
        try:
            if self.scan_job:
                self.root.after_cancel(self.scan_job)
        except Exception:
            pass
        self.scan_lock_files()

    def on_close(self):
        """关闭窗口前，安全停止定时任务"""
        try:
            if self.scan_job:
                self.root.after_cancel(self.scan_job)
        except Exception:
            pass
        self.root.destroy()

    def reset_monitor_state(self):
        """切换目录后重置监听状态，避免误报"""
        self.active_dwg_locks.clear()

    def scan_lock_files(self):
        """扫描当前目录下的图纸锁文件，识别“修改中”和“修改完成”状态"""
        root_dir = self.base_path.get().strip()
        current_locked = {}

        try:
            if root_dir and os.path.isdir(root_dir):
                for dir_path, _, file_names in os.walk(root_dir):
                    for file_name in file_names:
                        low_name = file_name.lower()
                        if low_name.endswith(".dwl") or low_name.endswith(".dwl2"):
                            lock_path = os.path.join(dir_path, file_name)
                            dwg_path = self.resolve_dwg_from_lock(lock_path)
                            if not dwg_path:
                                continue

                            current_locked[dwg_path] = lock_path
                            if dwg_path not in self.active_dwg_locks:
                                self.active_dwg_locks[dwg_path] = {
                                    "lock_path": lock_path,
                                    "version": self.extract_version_from_filename(dwg_path),
                                    "detected_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                }
                                logger.info("检测到图纸正在修改: %s", dwg_path)

                released_files = [dwg for dwg in list(self.active_dwg_locks.keys()) if dwg not in current_locked]
                for dwg_path in released_files:
                    lock_info = self.active_dwg_locks.pop(dwg_path, {})
                    self.handle_dwg_released(dwg_path, lock_info)

                self.monitor_status.set("自动监听：运行中（每3秒扫描）")
            else:
                self.monitor_status.set("自动监听：当前路径不可用")
        except Exception as e:
            logger.exception("扫描目录失败: %s", root_dir)
            self.monitor_status.set("自动监听：扫描异常，请检查目录")
        finally:
            self.scan_job = self.root.after(self.lock_poll_ms, self.scan_lock_files)

    def resolve_dwg_from_lock(self, lock_path):
        """根据 .dwl/.dwl2 锁文件反推对应的 .dwg 文件"""
        folder = os.path.dirname(lock_path)
        lock_name = os.path.basename(lock_path)
        base_name = re.sub(r'\.dwl2?$', '', lock_name, flags=re.IGNORECASE)
        target_name = f"{base_name}.dwg".lower()

        try:
            for file_name in os.listdir(folder):
                if file_name.lower() == target_name:
                    return os.path.join(folder, file_name)
        except Exception:
            logger.exception("解析锁文件对应图纸失败: %s", lock_path)
        return None

    def extract_version_from_filename(self, file_path):
        """从文件名末尾提取版本字段，如 A版、B版、1版"""
        raw_name = os.path.splitext(os.path.basename(file_path))[0]
        if "_" in raw_name:
            return raw_name.rsplit("_", 1)[1]
        return ""

    def handle_dwg_released(self, dwg_path, lock_info):
        """当检测到图纸编辑完成后，弹窗提示是否升级并写入修改记录"""
        finished_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        version_raw = lock_info.get("version") or self.extract_version_from_filename(dwg_path)
        version_display = version_raw or "未识别"
        file_name = os.path.basename(dwg_path)
        logger.info("检测到图纸修改完成: %s", dwg_path)

        need_upgrade = messagebox.askyesno(
            "检测到图纸已修改",
            f"检测到以下图纸已完成修改：\n\n{file_name}\n版本：{version_display}\n完成时间：{finished_time}\n\n是否立即升级为新版本？"
        )

        if need_upgrade:
            success, result = self.upgrade_file_by_path(dwg_path, input_version=version_raw)
            if success:
                messagebox.showinfo("升级完成", result)
            else:
                messagebox.showerror("升级失败", result)
        else:
            result = "用户取消升级"

        self.append_modify_record(dwg_path, finished_time, version_display, result)

    def append_modify_record(self, dwg_path, finished_time, version, result):
        """把修改记录写入图纸所在目录的“修改记录.xlsx”"""
        if Workbook is None or load_workbook is None:
            logger.error("未安装 openpyxl，无法生成修改记录.xlsx")
            messagebox.showwarning("记录失败", "当前环境缺少 openpyxl，无法生成修改记录.xlsx")
            return

        excel_path = os.path.join(os.path.dirname(dwg_path), "修改记录.xlsx")
        headers = ["修改文件", "文件路径", "修改时间", "版本", "处理结果"]
        col_widths = [28, 60, 22, 12, 28]

        try:
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
                ws = wb["修改记录"] if "修改记录" in wb.sheetnames else wb.active
                ws.title = "修改记录"
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "修改记录"

            first_row = [ws.cell(1, i).value for i in range(1, len(headers) + 1)]
            if first_row != headers:
                for idx, header in enumerate(headers, start=1):
                    ws.cell(1, idx, header)

            ws.append([
                os.path.basename(dwg_path),
                dwg_path,
                finished_time,
                version,
                result
            ])

            for idx, width in enumerate(col_widths, start=1):
                ws.column_dimensions[chr(64 + idx)].width = width

            wb.save(excel_path)
            logger.info("修改记录已写入: %s", excel_path)
        except PermissionError:
            logger.exception("修改记录文件被占用: %s", excel_path)
            messagebox.showwarning("记录失败", f"无法写入 {os.path.basename(excel_path)}，请先关闭已打开的 Excel 文件。")
        except Exception as e:
            logger.exception("写入修改记录失败: %s", dwg_path)
            messagebox.showerror("记录失败", f"写入修改记录失败：\n{str(e)}")
    
    def bulk_upgrade_files(self):
        """批量升级选中的所有 .dwg 和 .pdf 文件"""
        selected_ids = self.tree.selection()
        if not selected_ids:
            messagebox.showwarning("提示", "请先选择一个或多个文件")
            return

        success_count = 0
        skip_count = 0
        
        # 获取文本框输入的版本（用于判断是“统一改为某版本”还是“各自自动递增”）
        user_input_ver = self.version_input.get().strip()

        for item_id in selected_ids:
            old_path = self.get_full_path(item_id)
            
            # 过滤：只处理 dwg 和 pdf 文件
            if os.path.isdir(old_path) or not (old_path.lower().endswith(".dwg") or old_path.lower().endswith(".pdf")):
                skip_count += 1
                continue

            # --- 执行单个文件的升级逻辑 ---
            parent_dir = os.path.dirname(old_path)
            old_file_name = os.path.basename(old_path)
            raw_name, ext = os.path.splitext(old_file_name)
            
            # 提取当前文件自带的版本
            current_ver_in_file = raw_name.rsplit("_", 1)[1] if "_" in raw_name else ""
            pure_name = raw_name.rsplit("_", 1)[0] if "_" in raw_name else raw_name

            # 逻辑判定：
            # 如果文本框内容和当前文件版本一致，或者文本框为空 -> 各自自动递增
            # 如果文本框内容和当前文件版本不一致 -> 强制统一改为文本框的版本
            if user_input_ver == current_ver_in_file or not user_input_ver:
                final_ver = self.get_next_alpha_version(current_ver_in_file)
            else:
                final_ver = user_input_ver if "版" in user_input_ver else f"{user_input_ver}版"

            new_file_name = f"{pure_name}_{final_ver}{ext}"
            new_path = os.path.join(parent_dir, new_file_name)

            # 执行复制
            try:
                if not os.path.exists(new_path):
                    shutil.copy2(old_path, new_path)
                    success_count += 1
                else:
                    # 如果已存在则跳过，避免批量操作时弹出大量确认框
                    skip_count += 1
            except Exception:
                skip_count += 1

        self.refresh_tree()
        messagebox.showinfo("批量操作完成", f"成功升级: {success_count} 个文件\n跳过/失败: {skip_count} 个项目")

    def get_full_path(self, item_id):
        """辅助方法：根据 Treeview 的 item_id 溯源完整磁盘路径"""
        path_segments = []
        curr = item_id
        while curr:
            path_segments.insert(0, self.tree.item(curr, "text"))
            curr = self.tree.parent(curr)
        return os.path.join(self.base_path.get(), *path_segments)

    def show_context_menu(self, event):
        """右键选中树节点并弹出菜单"""
        item_id = self.tree.identify_row(event.y)
        if not item_id:
            return
        self.tree.selection_set(item_id)
        self.tree.focus(item_id)
        self.on_select(None)
        self.tree_menu.tk_popup(event.x_root, event.y_root)

    def open_selected_item(self):
        """打开当前选中的文件或文件夹"""
        target_path = self.current_selection.get()
        if not target_path:
            messagebox.showwarning("提示", "请先选择要打开的文件或文件夹")
            return
        if not os.path.exists(target_path):
            messagebox.showerror("打开失败", f"目标不存在：\n{target_path}")
            return
        try:
            os.startfile(target_path)
        except Exception as e:
            messagebox.showerror("打开失败", f"无法打开目标：\n{str(e)}")

    def open_selected_folder(self):
        """打开当前文件所在目录，或直接打开当前目录"""
        target_path = self.current_selection.get()
        if not target_path:
            messagebox.showwarning("提示", "请先选择一个项目")
            return

        folder_path = target_path if os.path.isdir(target_path) else os.path.dirname(target_path)
        if not os.path.exists(folder_path):
            messagebox.showerror("打开失败", f"目录不存在：\n{folder_path}")
            return
        try:
            os.startfile(folder_path)
        except Exception as e:
            messagebox.showerror("打开失败", f"无法打开目录：\n{str(e)}")
    
    def delete_items(self, event=None):
        """⭐ 新增功能：多选删除文件或文件夹"""
        selected_ids = self.tree.selection()
        if not selected_ids:
            messagebox.showwarning("提示", "请先选择要删除的项目")
            return

        # 确认提醒
        confirm = messagebox.askyesno(
            "彻底删除确认", 
            f"确定要删除选中的 {len(selected_ids)} 个项目吗？\n\n警告：此操作将从硬盘永久删除，不可撤销！"
        )
        
        if not confirm:
            return

        success_count = 0
        error_list = []

        for item_id in selected_ids:
            target_path = self.get_full_path(item_id)
            
            try:
                if os.path.isdir(target_path):
                    shutil.rmtree(target_path) # 删除整个文件夹
                elif os.path.isfile(target_path):
                    os.remove(target_path)     # 删除单个文件
                success_count += 1
            except Exception as e:
                error_list.append(f"{os.path.basename(target_path)}: {str(e)}")

        self.refresh_tree()
        
        if error_list:
            error_msg = "\n".join(error_list)
            messagebox.showwarning("部分删除失败", f"成功删除 {success_count} 项，失败 {len(error_list)} 项：\n{error_msg}")
        else:
            messagebox.showinfo("成功", f"已成功删除选中的 {success_count} 个项目")

    def on_select(self, event):
        selected = self.tree.selection()
        if not selected: return
        item_text = self.tree.item(selected[0], "text")
        
        path_segments = []
        curr = selected[0]
        while curr:
            path_segments.insert(0, self.tree.item(curr, "text"))
            curr = self.tree.parent(curr)
        
        full_path = os.path.join(self.base_path.get(), *path_segments)
        self.current_selection.set(full_path)

        if os.path.isdir(full_path):
            folder_version, folder_remark, _ = self.parse_version_folder_name(item_text)
            if folder_version:
                self.version_input.set(folder_version)
                self.remark_input.set(folder_remark)
            else:
                self.version_input.set(item_text)
                self.remark_input.set("")
        else:
            # ⭐ 同时支持 dwg 和 pdf 的版本解析
            low_text = item_text.lower()
            if "_" in item_text and (low_text.endswith(".dwg") or low_text.endswith(".pdf")):
                ver_part = item_text.split("_")[-1].rsplit(".", 1)[0]
                self.version_input.set(ver_part)
            else:
                self.version_input.set("")
            self.remark_input.set("")

    def parse_version_folder_name(self, folder_name):
        """解析版本文件夹名称，返回(版本号, 备注, 前缀)，如 V1.2.0-备注 -> ('1.2.0', '备注', 'V')"""
        pattern = re.compile(r'^(?P<prefix>[vV])(?P<version>\d+(?:\.\d+)+)(?:-(?P<remark>.*))?$')
        match = pattern.match(folder_name.strip())
        if not match:
            return "", "", ""
        version = match.group("version") or ""
        remark = (match.group("remark") or "").strip()
        prefix = match.group("prefix") or "V"
        return version, remark, prefix

    def is_valid_numeric_version(self, version_text):
        """校验版本号是否为纯数字段，如 1.2.0 或 1.2.0.1"""
        return bool(re.fullmatch(r'\d+(?:\.\d+)+', version_text.strip()))

    def normalize_version(self, version_text):
        """标准化版本号用于比较：去除前导零，保留段结构"""
        ver = version_text.strip()
        if ver.lower().startswith("v"):
            ver = ver[1:]
        if not self.is_valid_numeric_version(ver):
            return ""
        parts = [str(int(p)) for p in ver.split(".")]
        return ".".join(parts)

    def increment_numeric_version(self, version_text):
        """数字版本递增：只递增最后一段，如 1.2.0 -> 1.2.1"""
        parts = version_text.split(".")
        last_part = parts[-1]
        next_value = int(last_part) + 1
        # 保留末段位宽，便于兼容 001 这类格式
        parts[-1] = str(next_value).zfill(len(last_part)) if len(str(next_value)) <= len(last_part) else str(next_value)
        return ".".join(parts)

    def build_version_folder_name(self, version_text, remark_text, prefix="V"):
        """按规则构建版本文件夹名：V1.2.0.1-备注"""
        safe_remark = remark_text.strip().replace("\\", "_").replace("/", "_").replace(":", "_").replace("*", "_")
        safe_remark = safe_remark.replace("?", "_").replace('"', "_").replace("<", "_").replace(">", "_").replace("|", "_")
        base_name = f"{prefix}{version_text}"
        return f"{base_name}-{safe_remark}" if safe_remark else base_name

    def resolve_version_folder_parent(self):
        """确定版本文件夹操作的目标父目录"""
        selected_path = self.current_selection.get()
        if selected_path and os.path.isdir(selected_path):
            return os.path.dirname(selected_path)
        return self.base_path.get().strip()

    def get_next_alpha_version(self, current_ver):
        """字母版本递增算法: A->B, ..., Z->AA, AA->AB, ..., AZ->BA, ..., ZZ->AAA"""
        pure_ver = current_ver.replace("版", "").strip().upper()
        if not pure_ver:
            return "B版"
        # 数字版本递增
        if pure_ver.isdigit():
            return f"{int(pure_ver) + 1}版"
        # 字母版本递增
        if pure_ver.isalpha():
            # 从右向左找到第一个不是Z的字符
            chars = list(pure_ver)
            i = len(chars) - 1
            while i >= 0 and chars[i] == 'Z':
                chars[i] = 'A'
                i -= 1
            if i < 0:
                # 全部是Z，进位增加一位（如 Z->AA, ZZ->AAA）
                return 'A' * (len(chars) + 1) + '版'
            else:
                # 当前位递增
                chars[i] = chr(ord(chars[i]) + 1)
                return ''.join(chars) + '版'
        # 混合版本，降级处理
        return f"{pure_ver}.1版"

    def upgrade_file_by_path(self, old_path, input_version=None):
        """按指定路径升级文件，便于手动操作和自动监听共用"""
        if not old_path or not (old_path.lower().endswith(".dwg") or old_path.lower().endswith(".pdf")):
            return False, "请选择一个 .dwg 或 .pdf 文件"

        parent_dir = os.path.dirname(old_path)
        old_file_name = os.path.basename(old_path)
        raw_name, ext = os.path.splitext(old_file_name)

        current_ver_in_file = raw_name.rsplit("_", 1)[1] if "_" in raw_name else ""
        pure_name = raw_name.rsplit("_", 1)[0] if "_" in raw_name else raw_name
        user_input_ver = input_version.strip() if isinstance(input_version, str) else self.version_input.get().strip()

        if user_input_ver == current_ver_in_file or not user_input_ver:
            final_ver = self.get_next_alpha_version(current_ver_in_file)
        else:
            final_ver = user_input_ver if "版" in user_input_ver else f"{user_input_ver}版"

        new_file_name = f"{pure_name}_{final_ver}{ext}"
        new_path = os.path.join(parent_dir, new_file_name)

        if os.path.exists(new_path):
            if not messagebox.askyesno("确认", f"文件 {new_file_name} 已存在，是否覆盖？"):
                return False, "用户取消覆盖"

        try:
            shutil.copy2(old_path, new_path)
            logger.info("文件升级成功: %s -> %s", old_path, new_path)
            self.refresh_tree()
            return True, f"已生成新版本文件：\n{new_file_name}"
        except Exception as e:
            logger.exception("文件升级失败: %s", old_path)
            return False, str(e)

    def upgrade_file(self, event=None):
        """升级文件逻辑：支持 .dwg 和 .pdf"""
        success, result = self.upgrade_file_by_path(self.current_selection.get())
        if success:
            messagebox.showinfo("成功", result)
        else:
            messagebox.showwarning("提示", result)

    # --- 辅助方法 ---
    def browse_path(self):
        p = filedialog.askdirectory()
        if p:
            self.base_path.set(p)
            self.reset_monitor_state()
            self.refresh_tree()

    def get_mtime(self, p):
        return datetime.fromtimestamp(os.path.getmtime(p)).strftime('%Y-%m-%d %H:%M')

    def refresh_tree(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        root_dir = self.base_path.get()
        if not os.path.exists(root_dir): return
        v_pattern = re.compile(r'^[vV]\d+(?:\.\d+)+(?:-.*)?$')
        try:
            for entry in sorted(os.scandir(root_dir), key=lambda e: e.name.lower()):
                if entry.is_dir() and v_pattern.match(entry.name):
                    node = self.tree.insert("", "end", text=entry.name, values=(self.get_mtime(entry.path), "版本根目录"))
                    self._fill_tree(node, entry.path)
        except: pass

    def _fill_tree(self, parent, path):
        """填充树状图，并排除 .bak 文件"""
        try:
            # 排序：文件夹在前，文件在后
            entries = sorted(os.scandir(path), key=lambda e: (not e.is_dir(), e.name.lower()))
            for entry in entries:
                # ⭐ 关键修改：如果是文件且后缀是 .bak，则直接跳过
                if entry.is_file() and entry.name.lower().endswith(".bak"):
                    continue

                ext_str = os.path.splitext(entry.name)[1].upper() if entry.is_file() else "文件夹"
                node = self.tree.insert(parent, "end", text=entry.name, values=(self.get_mtime(entry.path), ext_str))
                
                if entry.is_dir():
                    self._fill_tree(node, entry.path)
        except:
            pass

    def upgrade_folder(self, event=None):
        old_path = self.current_selection.get()
        if not old_path or not os.path.isdir(old_path):
            messagebox.showwarning("提示", "请先选择一个版本文件夹")
            return

        old_name = os.path.basename(old_path)
        current_version, current_remark, prefix = self.parse_version_folder_name(old_name)
        if not current_version:
            messagebox.showwarning("提示", "选中的文件夹名称不符合规则：V1.2.0.1-备注")
            return

        input_version = self.version_input.get().strip()
        input_remark = self.remark_input.get().strip()
        if not input_version:
            input_version = current_version

        # 允许输入带 V/v 前缀
        if input_version.lower().startswith("v"):
            input_version = input_version[1:]

        if not self.is_valid_numeric_version(input_version):
            messagebox.showwarning("提示", "版本号格式无效，请输入如 1.2.0 或 1.2.0.1")
            return

        if self.normalize_version(input_version) == self.normalize_version(current_version):
            new_version = self.increment_numeric_version(current_version)
        else:
            new_version = input_version

        final_remark = input_remark if input_remark else current_remark
        new_name = self.build_version_folder_name(new_version, final_remark, prefix=prefix or "V")
        new_path = os.path.join(os.path.dirname(old_path), new_name)

        if os.path.exists(new_path):
            messagebox.showwarning("提示", f"目标文件夹已存在：\n{new_name}")
            return

        try:
            shutil.copytree(old_path, new_path)
            logger.info("文件夹升级成功: %s -> %s", old_path, new_path)
            self.version_input.set(new_version)
            self.remark_input.set(final_remark)
            self.refresh_tree()
            messagebox.showinfo("成功", f"已升级为新版本文件夹：\n{new_name}")
        except Exception as e:
            logger.exception("文件夹升级失败: %s", old_path)
            messagebox.showerror("失败", str(e))

    def create_version_folder(self):
        """按输入的版本号和备注创建新的空版本文件夹"""
        parent_dir = self.resolve_version_folder_parent()
        if not parent_dir or not os.path.isdir(parent_dir):
            messagebox.showwarning("提示", "当前目标路径不可用，请先选择有效目录")
            return

        input_version = self.version_input.get().strip()
        input_remark = self.remark_input.get().strip()
        if not input_version:
            messagebox.showwarning("提示", "请先输入版本号，例如 1.2.0.1")
            return

        if input_version.lower().startswith("v"):
            prefix = input_version[0]
            input_version = input_version[1:]
        else:
            prefix = "V"

        if not self.is_valid_numeric_version(input_version):
            messagebox.showwarning("提示", "版本号格式无效，请输入如 1.2.0 或 1.2.0.1")
            return

        new_name = self.build_version_folder_name(input_version, input_remark, prefix=prefix)
        new_path = os.path.join(parent_dir, new_name)
        if os.path.exists(new_path):
            messagebox.showwarning("提示", f"文件夹已存在：\n{new_name}")
            return

        try:
            os.makedirs(new_path, exist_ok=False)
            logger.info("新建版本文件夹成功: %s", new_path)
            self.refresh_tree()
            messagebox.showinfo("成功", f"已新建版本文件夹：\n{new_name}")
        except Exception as e:
            logger.exception("新建版本文件夹失败: %s", new_path)
            messagebox.showerror("失败", str(e))

    def parse_version(self, ver_str):
        """解析版本号，返回可比较的值
        支持任意长度字母版本：A=1, Z=26, AA=27, AZ=52, BA=53, ..., ZZ=702, AAA=703...
        """
        ver_str = ver_str.replace("版", "").strip().upper()
        if not ver_str:
            return (0, 0)
        # 纯数字版本
        if ver_str.isdigit():
            return (1, int(ver_str))
        # 字母版本（类似Excel列号算法）
        if ver_str.isalpha():
            result = 0
            for char in ver_str:
                result = result * 26 + (ord(char) - ord('A') + 1)
            return (2, result)
        # 混合版本，按字符串比较
        return (3, ver_str)

    def delete_old_versions(self):
        """一键删除选中文件的旧版本，只保留最新版"""
        selected_ids = self.tree.selection()
        if not selected_ids:
            messagebox.showwarning("提示", "请先选择一个或多个文件")
            return

        # 1. 过滤出 dwg 和 pdf 文件
        target_files = []
        for item_id in selected_ids:
            file_path = self.get_full_path(item_id)
            if os.path.isfile(file_path):
                ext = file_path.lower().split('.')[-1]
                if ext in ['dwg', 'pdf']:
                    target_files.append(file_path)

        if not target_files:
            messagebox.showwarning("提示", "选中的项目中没有 .dwg 或 .pdf 文件")
            return

        # 2. 按纯名称分组（去掉版本后缀）
        file_groups = {}  # {纯名称: [(版本号, 文件路径), ...]}
        for file_path in target_files:
            file_name = os.path.basename(file_path)
            raw_name, ext = os.path.splitext(file_name)
            
            if "_" in raw_name:
                pure_name = raw_name.rsplit("_", 1)[0]
                ver_part = raw_name.rsplit("_", 1)[1]
            else:
                pure_name = raw_name
                ver_part = ""
            
            key = f"{pure_name}{ext}"
            if key not in file_groups:
                file_groups[key] = []
            file_groups[key].append((ver_part, file_path))

        # 3. 对每组文件比较版本号，找出最新版本
        files_to_delete = []
        files_to_keep = []
        
        for group_name, versions in file_groups.items():
            if len(versions) <= 1:
                continue  # 只有一个版本，不需要删除
            
            # 按版本号排序，找出最新版本
            sorted_versions = sorted(versions, key=lambda x: self.parse_version(x[0]), reverse=True)
            
            # 保留最新版本
            files_to_keep.append(sorted_versions[0][1])
            # 删除其他旧版本
            for ver, file_path in sorted_versions[1:]:
                files_to_delete.append(file_path)

        if not files_to_delete:
            messagebox.showinfo("提示", "选中的文件中没有需要删除的旧版本")
            return

        # 4. 显示确认对话框
        delete_list = "\n".join([os.path.basename(f) for f in files_to_delete])
        keep_list = "\n".join([os.path.basename(f) for f in files_to_keep])
        
        confirm_msg = f"将保留以下最新版本:\n{keep_list}\n\n将删除以下旧版本:\n{delete_list}\n\n确定要删除这些旧版本吗？"
        confirm = messagebox.askyesno("确认删除旧版本", confirm_msg)
        
        if not confirm:
            return

        # 5. 执行删除
        success_count = 0
        error_list = []
        
        for file_path in files_to_delete:
            try:
                os.remove(file_path)
                success_count += 1
            except Exception as e:
                error_list.append(f"{os.path.basename(file_path)}: {str(e)}")

        self.refresh_tree()
        
        if error_list:
            error_msg = "\n".join(error_list)
            messagebox.showwarning("部分删除失败", 
                f"成功删除 {success_count} 个旧版本，失败 {len(error_list)} 个：\n{error_msg}")
        else:
            messagebox.showinfo("操作完成", 
                f"已成功删除 {success_count} 个旧版本\n保留了 {len(files_to_keep)} 个最新版本")

if __name__ == "__main__":
    root = tk.Tk()
    
    # --- 核心修改：获取命令行参数 ---
    # 如果通过右键菜单启动，sys.argv[1] 会包含文件夹路径
    if len(sys.argv) > 1:
        initial_path = sys.argv[1]
    else:
        # 如果直接双击运行，则默认打开程序所在目录或指定路径
        initial_path = os.getcwd() 

    app = DrawingManager(root, default_path=initial_path)
    root.mainloop()